"""
Export OMR results to Excel (.xlsx) and CSV.
Layout per client spec:
  Columns: Participante | Pág | Card | Jogo 1 | ... | Jogo 8 | Confiança | Total
  Each card = one row.
"""

import csv
import io
from pathlib import Path
from typing import List, Dict, Optional

from . import database as db_mod
from .omr.recognizer import OPTION_LABELS


GAME_HEADERS = [f"Jogo {i+1}" for i in range(8)]


def _choice_text(choice: Optional[int]) -> str:
    if choice is None:
        return "?"
    try:
        return OPTION_LABELS[choice]
    except IndexError:
        return "?"


def _build_rows(session_id: int, db_path: Path = None) -> List[Dict]:
    """Assemble one flat row per card with all game choices."""
    cards = db_mod.get_cards(session_id, db_path)
    official = db_mod.get_official_results(session_id, db_path)
    ranking_map: Dict[int, int] = {}
    for r in db_mod.get_ranking(session_id, db_path):
        ranking_map[r["card_id"]] = r["score"]

    # Batch-load all marks once (single query) instead of one query per card
    marks_by_card = db_mod.get_marks_for_session(session_id, db_path)

    rows = []
    for card in cards:
        marks = marks_by_card.get(card["id"], [])
        marks_by_game = {m["game"]: m for m in marks}

        choices = []
        low_conf = False
        for g in range(8):
            m = marks_by_game.get(g)
            if m:
                choices.append(_choice_text(m["choice"]))
                if m["needs_review"]:
                    low_conf = True
            else:
                choices.append("?")
                low_conf = True

        participant = card.get("participant") or f"Pág {card['page']} #{card['card_index']+1}"
        score = ranking_map.get(card["id"], "")
        rows.append({
            "Participante": participant,
            "Nome": card.get("nome") or "",   # nome do apostador (cartela digital) — p/ o site
            "Página": card["page"],
            "Card": card["card_index"] + 1,
            **{f"Jogo {i+1}": choices[i] for i in range(8)},
            "Revisão": "SIM" if low_conf else "",
            "Total": score,
        })

    return rows


def export_excel(session_id: int, output_path: str,
                 db_path: Path = None) -> None:
    try:
        import openpyxl
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    rows = _build_rows(session_id, db_path)
    session = db_mod.get_session(session_id, db_path)
    session_name = session["name"] if session else f"Sessão {session_id}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado"

    # Header styles
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    review_fill = PatternFill("solid", fgColor="FFD700")
    alt_fill    = PatternFill("solid", fgColor="EEF4FB")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Participante", "Nome", "Página", "Card"] + \
              [f"Jogo {i+1}" for i in range(8)] + ["Revisão", "Total"]

    # Write header row
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        is_alt = row_idx % 2 == 0
        for col, h in enumerate(headers, start=1):
            val = row.get(h, "")
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.border = border
            if row.get("Revisão") == "SIM" and h == "Revisão":
                cell.fill = review_fill
                cell.font = Font(bold=True)
            elif is_alt:
                cell.fill = alt_fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 25)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add ranking sheet if scores exist
    ranking = db_mod.get_ranking(session_id, db_path)
    if ranking:
        ws2 = wb.create_sheet(title="Ranking")
        rank_headers = ["Posição", "Participante", "Acertos"]
        for col, h in enumerate(rank_headers, start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        pos = 0
        prev_score = None
        for i, r in enumerate(ranking):
            if r["score"] != prev_score:
                pos = i + 1
                prev_score = r["score"]
            label = r.get("participant") or f"Pág {r['page']} #{r['card_index']+1}"
            for col, val in enumerate([pos, label, r["score"]], start=1):
                cell = ws2.cell(row=i + 2, column=col, value=val)
                cell.alignment = center
                cell.border = border
                if i % 2 == 0:
                    cell.fill = alt_fill

        for col in ws2.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 30)

    wb.save(output_path)


def export_csv(session_id: int, output_path: str,
               db_path: Path = None) -> None:
    rows = _build_rows(session_id, db_path)
    if not rows:
        return
    headers = list(rows[0].keys())
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
